"""
Excel Column Value Mover

This script processes an Excel file and moves values from column C to column D
based on specific conditions:
- Starting from row 9
- For every row where column B has a value (target row)
- Check ONLY the immediate next row in column C
- If there is a value in column C of the immediate next row, move it to column D of the target row
- If column C of the immediate next row is empty, move on to the next target row
"""

import openpyxl
import sys
import os
from pathlib import Path

# Try to import xlrd for .xls file support
try:
    import xlrd
    from openpyxl import Workbook
    XLS_SUPPORT = True
except ImportError:
    XLS_SUPPORT = False


def convert_xls_to_xlsx(xls_path, xlsx_path):
    """Convert .xls file to .xlsx format."""
    if not XLS_SUPPORT:
        raise ImportError("xlrd library is required for .xls files. Install it with: pip install xlrd")
    
    # Read .xls file
    xls_workbook = xlrd.open_workbook(xls_path)
    xls_sheet = xls_workbook.sheet_by_index(0)
    
    # Create new .xlsx workbook
    xlsx_workbook = Workbook()
    xlsx_worksheet = xlsx_workbook.active
    
    # Copy data from .xls to .xlsx
    for row_idx in range(xls_sheet.nrows):
        for col_idx in range(xls_sheet.ncols):
            cell_value = xls_sheet.cell_value(row_idx, col_idx)
            xlsx_worksheet.cell(row=row_idx + 1, column=col_idx + 1, value=cell_value)
    
    # Save as .xlsx
    xlsx_workbook.save(xlsx_path)
    return xlsx_path


def move_column_values(file_path, output_path=None):
    """
    Process Excel file and move column C values to column D based on conditions.
    
    Args:
        file_path (str): Path to the input Excel file (.xlsx or .xls)
        output_path (str, optional): Path to save the output file. 
                                     If None, overwrites the original file.
    """
    try:
        # Normalize the path (handle spaces and absolute paths)
        file_path = os.path.normpath(file_path)
        
        # Handle .xls files by converting to .xlsx first
        temp_xlsx_path = None
        if file_path.lower().endswith('.xls'):
            if not XLS_SUPPORT:
                print("Error: .xls files require xlrd library.")
                print("Install it with: pip install xlrd")
                sys.exit(1)
            
            # Create temporary .xlsx file
            temp_xlsx_path = file_path.rsplit('.', 1)[0] + '_temp.xlsx'
            print(f"Converting .xls to .xlsx format...")
            convert_xls_to_xlsx(file_path, temp_xlsx_path)
            file_path = temp_xlsx_path
        
        # Load the workbook
        workbook = openpyxl.load_workbook(file_path)
        worksheet = workbook.active
        
        # Get the maximum row to avoid going out of bounds
        max_row = worksheet.max_row
        
        # Start from row 9 (1-indexed)
        start_row = 9
        
        # Process each row starting from row 9
        for target_row in range(start_row, max_row + 1):
            # Check if column B (index 2) has a value in the target row
            cell_b = worksheet.cell(row=target_row, column=2)
            
            if cell_b.value is not None and str(cell_b.value).strip() != '':
                # This is a target row - check ONLY the immediate next row in column C
                check_row = target_row + 1
                
                # Only check if the next row exists
                if check_row <= max_row:
                    cell_c = worksheet.cell(row=check_row, column=3)
                    
                    # If there is a value in column C of the immediate next row, move it
                    if cell_c.value is not None and str(cell_c.value).strip() != '':
                        value_to_move = cell_c.value
                        
                        # Set the value in column D of the target row
                        cell_d = worksheet.cell(row=target_row, column=4)
                        cell_d.value = value_to_move
                        
                        # Clear the original value in column C (move, not copy)
                        worksheet.cell(row=check_row, column=3).value = None
                        
                        print(f"Row {target_row}: Moved value '{value_to_move}' from C{check_row} to D{target_row}")
                    # If empty, we just move on to the next target row (no action needed)
        
        # Determine output path
        if output_path is None:
            # If original was .xls, save as .xlsx with same name
            original_path = sys.argv[1] if len(sys.argv) > 1 else file_path
            original_path = os.path.normpath(original_path)
            if original_path.lower().endswith('.xls'):
                output_path = original_path.rsplit('.', 1)[0] + '.xlsx'
            else:
                output_path = file_path
        else:
            output_path = os.path.normpath(output_path)
        
        # Ensure output directory exists
        output_dir = os.path.dirname(output_path)
        if output_dir and not os.path.exists(output_dir):
            os.makedirs(output_dir)
        
        workbook.save(output_path)
        print(f"\nProcessing complete! File saved to: {output_path}")
        
        # Clean up temporary file if created
        if temp_xlsx_path and os.path.exists(temp_xlsx_path):
            os.remove(temp_xlsx_path)
        
    except FileNotFoundError:
        print(f"Error: File '{file_path}' not found.")
        sys.exit(1)
    except ImportError as e:
        print(f"Error: {str(e)}")
        sys.exit(1)
    except Exception as e:
        print(f"Error processing file: {str(e)}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


def main():
    """Main function to handle command line arguments."""
    if len(sys.argv) < 2:
        print("Usage: python move_column_values.py <input_file> [output_file]")
        print("\nExamples:")
        print("  python move_column_values.py data.xlsx")
        print("  python move_column_values.py data.xlsx output.xlsx")
        print('  python move_column_values.py "C:\\Path\\With Spaces\\file.xlsx"')
        print('  python move_column_values.py "C:\\Users\\brian\\OneDrive\\Desktop\\United Panel\\Excel Files for Importing\\Supplier\\ucs creditor list.xls"')
        sys.exit(1)
    
    input_file = sys.argv[1]
    output_file = sys.argv[2] if len(sys.argv) > 2 else None
    
    # Normalize the path to handle Windows paths properly
    input_file = os.path.normpath(input_file)
    
    # Check if input file exists
    if not os.path.exists(input_file):
        print(f"Error: Input file '{input_file}' does not exist.")
        print(f"Please check the path and try again.")
        sys.exit(1)
    
    # Process the file
    move_column_values(input_file, output_file)


if __name__ == "__main__":
    main()

